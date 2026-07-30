"""Excel 导出测试。"""
import io

from openpyxl import load_workbook

from research_agent.analysis.export import build_workbook
from research_agent.survey import build_survey


def test_export_workbook():
    survey = build_survey(
        "满意度",
        [
            {"type": "radio", "title": "满意吗", "options": ["满意", "不满意"]},
            {"type": "radio-nps", "title": "推荐", "min": 0, "max": 10},
            {"type": "textarea", "title": "建议"},
        ],
    )
    qs = survey.questions
    fr, fn, ft = qs[0].field, qs[1].field, qs[2].field
    opt = {o.text: o.hash for o in qs[0].options}
    rows = [
        {"data": {fr: opt["满意"], fn: 10, ft: "很好"}, "created_at": "2026-07-29"},
        {"data": {fr: opt["不满意"], fn: 3}, "created_at": "2026-07-29"},
    ]

    content = build_workbook(survey.model_dump(), "满意度", rows)
    assert content[:2] == b"PK"  # xlsx 本质是 zip

    wb = load_workbook(io.BytesIO(content))
    assert wb.sheetnames == ["答卷明细", "分题统计"]
    ws = wb["答卷明细"]
    assert ws.cell(1, 1).value == "满意吗"      # 表头
    assert ws.cell(2, 1).value == "满意"         # 第1份答卷第1题（hash 已还原为文案）
    assert ws.cell(3, 1).value == "不满意"
